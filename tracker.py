
import pandas as pd
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
import requests
import json
import re
import gspread
import openpyxl
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

#gets the google sheet with member nation names and discord ID (not ness with a proper backend)
gaccount = gspread.service_account(filename = 'war pig-9478580af5de.json')
gsheet = gaccount.open("Discord Tracking Sheet").sheet1

API_KEY = '69e9cc72114cd2'

def load_wars():
	'''
	Gets list of wars in an alliance that are active
	'''
	war_json = requests.get(f'https://politicsandwar.com/api/wars/300&alliance_id=7452&key={API_KEY}').json()
	
	#Makes sure not out of API keys
	if war_json['success']:
		war_list = war_json['wars']
		member_records = {}
		#Dataframe to know where the war left off, so it doesn't double count 
		next_save = pd.DataFrame(columns = ['War ID', 'Last War Attack ID'])
		#Grabs each ear
		for war in war_list:
			save = last_save(war['warID'])
			#Checks if war is still active
			if war["status"] == "Active" or war["status"] == 'Defender Offered Peace' or war["status"] == 'Attacker Offered Peace' or last_save("warID") != None:
				#Determines if member is attacker or defender
				if war['attackerAA'] == 'Children of the Light':
					records = calc_war(war['attackerID'], war['warID'], save)
					#Checks if member is already in dictionary
					if war['attackerID'] in member_records:
						for key in member_records[war['attackerID']][0]:
							member_records[war['attackerID']][0][key] += records[0][key]
						for key in member_records[war['attackerID']][1]:
							member_records[war['attackerID']][1][key] += records[1][key]
					#Creates member if not in dictionary
					else:
						member_records[war['attackerID']] = records[0:2]
					if records[2][0]:
						next_save.loc[len(next_save.index)] = [war['warID'], records[2][1]]
				#If defender
				elif war['defenderAA'] == 'Children of the Light':
					records = calc_war(war['defenderID'], war['warID'], save)
					if war['defenderID'] in member_records:
						#member_records[war['defenderID']] += records[0:2]
						for key in member_records[war['defenderID']][0]:
							member_records[war['defenderID']][0][key] += records[0][key]
						for key in member_records[war['defenderID']][1]:
							member_records[war['defenderID']][1][key] += records[1][key]
					else:
						member_records[war['defenderID']] = records[0:2]
					if records[2][0]:
						next_save.loc[len(next_save.index)] = [war['warID'], records[2][1]]
		#Saves to excel to know last war attack that occured
		next_save.to_excel('Last Active Wars.xlsx', index = False)
		update_stats(member_records)




air_attacks = {'airstrike2': 'sold', 'airstrike3': 'tanks', 'airstrike4': 'mon_destroyed', 'airstrike5': 'ships'}
def calc_war(member_id, war_id, min_attack_id = None):
	'''
	Calculates the war damage
	:param member_id: the id of the alliance member
	:param war_id: the id of the war
	:param min_attack_id: the last war attack registered
	:return damage_dealt: damage member dealt this war
	:return damage_taken: damage member took this war
	:return war_active: whether if war is still active and needs to be scrapped in future

	'''

	#API to get war casualities and everything
	calc_json = requests.get(f'https://politicsandwar.com/api/war-attacks/key={API_KEY}&war_id={war_id}&min_war_attack_id={min_attack_id}').json()
	if calc_json['success']:
		member_dmg_dealt = {"infra_destroyed": 0, 'money_looted': 0, "infra_destroyed_value": 0, "gas": 0, "mun": 0,
		"sold": 0, "tanks": 0, "air": 0, "ships": 0, 'mon_destroyed': 0, 'nuke': 0, 'miss': 0, 'beige_loot': [
		0,0,0,0,0,0,0,0,0,0,0]}

		member_dmg_taken = {"infra_destroyed": 0, 'money_looted': 0, "infra_destroyed_value": 0, "gas": 0, "mun": 0,
		"sold": 0, "tanks": 0, "air": 0, "ships": 0, 'mon_destroyed': 0, 'nuke': 0, 'miss': 0, 'beige_loot': [
		0,0,0,0,0,0,0,0,0,0,0]}

		active_war = [True, None]

		#If the war has war attacks
		if len(calc_json['war_attacks']) > 0:
			active_war = [True, calc_json['war_attacks'][0]["war_attack_id"]]

		#No one has attacked yet
		else: 
			return [member_dmg_dealt, member_dmg_taken, active_war]

		#Checks how many war attacks to process
		attacks_to_process = -1
		if min_attack_id == None:
			attacks_to_process = len(calc_json['war_attacks'])

		#Loops through war attacks and adds them up for everything used and killed
		for war in calc_json['war_attacks'][:attacks_to_process]:
			#If member is attacker
			if int(war['attacker_nation_id']) == member_id:
				member_dmg_dealt['mun'] += float(war['def_mun_used'])
				member_dmg_dealt['gas'] += float(war['def_gas_used'])
				member_dmg_taken['mun'] += float(war['att_mun_used'])
				member_dmg_taken['gas'] += float(war['att_gas_used'])
				member_dmg_dealt['infra_destroyed'] += float(war['infra_destroyed'])
				member_dmg_dealt['money_looted'] += float(war['money_looted'])
				member_dmg_dealt['infra_destroyed_value'] += float(war['infra_destroyed_value'])
				#If it is an airstrike
				if re.search(r'airstrike', war['attack_type']):
					member_dmg_dealt['air'] += float(war['defcas1'])
					member_dmg_taken['air'] += float(war['attcas1'])
					#If it is not an air to air or an air to infra
					if war['attack_type'] != 'airstrike1' and war['attack_type'] != 'airstrike6':
						member_dmg_dealt[air_attacks[war['attack_type']]] += float(war['defcas2'])
				#If it is a ground attack
				elif war['attack_type'] == 'ground':
					member_dmg_dealt['sold'] += float(war['defcas1'])
					member_dmg_dealt['tanks'] += float(war['defcas2'])
					member_dmg_taken['sold'] += float(war['attcas1'])
					member_dmg_taken['tanks'] += float(war['attcas2'])
					member_dmg_dealt['air'] += float(war['aircraft_killed_by_tanks'])
				#If it is naval
				elif war['attack_type'] == 'naval':
					member_dmg_dealt['ships'] += float(war['defcas1'])
					member_dmg_taken['ships'] += float(war['attcas1'])

				elif war['attack_type'] == 'nuke' or war['attack_type'] == 'nukef':
					member_dmg_taken['nuke'] += 1

				elif war['attack_type'] == 'missile' or war['attack_type'] == 'missilef':
					member_dmg_taken['miss'] += 1

				#If it is a victory, gets the beige loot
				elif war['attack_type'] == 'victory':
					loot = war['note'].split('won the war and looted $')[1]
					loot = re.sub(r"[^.0-9\\ ]", '', loot)
					while '  ' in loot:
						loot = loot.replace('  ', ' ')
					loot = loot.split(' ')
					for index, rss in enumerate(loot[1:12]):
						member_dmg_dealt['beige_loot'][index] += float(rss)

				#If it is victory, grab looted from alliance
				elif war['attack_type'] == 'a_loot':
					loot = war['note'].split('alliance bank, taking: $')[1]
					loot = re.sub(r"[^.0-9\\ ]", '', loot)
					while '  ' in loot:
						loot = loot.replace('  ', ' ')
					loot = loot.split(' ')
					for index, rss in enumerate(loot[1:12]):
						member_dmg_dealt['beige_loot'][index] += float(rss)
					active_war[0] = False

			#If the member is not the attacker
			else:
				member_dmg_taken['mun'] += float(war['def_mun_used'])
				member_dmg_taken['gas'] += float(war['def_gas_used'])
				member_dmg_dealt['mun'] += float(war['att_mun_used'])
				member_dmg_dealt['gas'] += float(war['att_gas_used'])
				member_dmg_taken['infra_destroyed'] += float(war['infra_destroyed'])
				member_dmg_taken['money_looted'] += float(war['money_looted'])
				member_dmg_taken['infra_destroyed_value'] += float(war['infra_destroyed_value'])
				if re.search(r'airstrike', war['attack_type']):
					member_dmg_taken['air'] += float(war['defcas1'])
					member_dmg_dealt['air'] += float(war['attcas1'])
					if war['attack_type'] != 'airstrike1' and war['attack_type'] != 'airstrike6':
						member_dmg_taken[air_attacks[war['attack_type']]] += float(war['defcas2'])

				elif war['attack_type'] == 'ground':
					member_dmg_taken['sold'] += float(war['defcas1'])
					member_dmg_taken['tanks'] += float(war['defcas2'])
					member_dmg_dealt['sold'] += float(war['attcas1'])
					member_dmg_dealt['tanks'] += float(war['attcas2'])
					member_dmg_taken['air'] += float(war['aircraft_killed_by_tanks'])

				elif war['attack_type'] == 'naval':
					member_dmg_taken['ships'] += float(war['defcas1'])
					member_dmg_dealt['ships'] += float(war['attcas1'])

				elif war['attack_type'] == 'nuke' or war['attack_type'] == 'nukef':
					member_dmg_dealt['nuke'] += 1

				elif war['attack_type'] == 'missile' or war['attack_type'] == 'missilef':
					member_dmg_dealt['miss'] += 1

				elif war['attack_type'] == 'victory':
					loot = war['note'].split('won the war and looted $')[1]
					loot = re.sub(r"[^.0-9\\ ]", '', loot)
					while '  ' in loot:
						loot = loot.replace('  ', ' ')
					loot = loot.split(' ')
					for index, rss in enumerate(loot[1:12]):
						member_dmg_taken['beige_loot'][index] += float(rss)

				elif war['attack_type'] == 'a_loot':
					loot = war['note'].split('alliance bank, taking: $')[1]
					loot = re.sub(r"[^.0-9\\ ]", '', loot)
					while '  ' in loot:
						loot = loot.replace('  ', ' ')
					loot = loot.split(' ')
					for index, rss in enumerate(loot[1:12]):
						member_dmg_taken['beige_loot'][index] += float(rss)
					active_war[0] = False

		return [member_dmg_dealt, member_dmg_taken, active_war]

def last_save(war_id):
	try:
		print(war_id)
		print(last_war_save[last_war_save['War ID'] == war_id]['Last War Attack ID'].item())
		return last_war_save[last_war_save['War ID'] == war_id]['Last War Attack ID'].item()
	except:
		return None



def update_stats(member_records):
	'''
	Updates the stats in the spreadsheet
	:member_records: the new war stats to add
	'''

	rss_prices = {}
	#Gets current rss prices
	for resource in ['Coal', 'Oil', 'Uranium', 'Iron', 'Bauxite', 'Lead', 'Gasoline', 'Munitions', 'Steel', 'Aluminum', 'Food']:
		rss_json = war_json = requests.get(f'https://politicsandwar.com/api/tradeprice/resource={resource.lower()}&key={API_KEY}').json()
		rss_prices[resource] = int(rss_json['avgprice'])

	#Loads spreadsheet and member list from google spreadsheet
	book = load_workbook('Stat Tracker.xlsx')
	member_names = gsheet.col_values(1)
	member_names.pop(0)
	nation_id = gsheet.col_values(2)
	nation_id.pop(0)
	member_dict = dict(zip(nation_id, member_names))

	#Goes through each member and updates
	for member in member_records:
		sheet = book.active
		#Creates new sheet if that member wasnt in stat tracker before
		if member_dict.get(str(member), str(member)) not in book.sheetnames:
			sheet = book.create_sheet(member_dict.get(str(member), str(member)))
			sheet['A1'] = member_dict.get(str(member), str(member))
			sheet['A1'].font = Font(size = 20)
			names = ['Infrastructure','Soldiers','Tanks','Planes','Ships','Missiles','Nukes','All Units','Gas and Munition','Money','Looted','Total']
			
			for i, name in enumerate(names):
				sheet.cell(row = 2, column = i+2).value = name
				sheet.cell(row = 2, column = i+2).fill = PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid')
				sheet.cell(row = 2, column = i+2).font = Font(color = 'FFFFFF')

				sheet.cell(row = 21, column = i+2).value = name
				sheet.cell(row = 21, column = i+2).fill = PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid')
				sheet.cell(row = 21, column = i+2).font = Font(color = 'FFFFFF')

				sheet.cell(row = 4, column = i+1).fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
			sheet.cell(row = 4, column = 13).fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
			sheet.cell(row = 3, column = 1).value = 'Total Net Damage'
			sheet.cell(row = 4, column = 1).value = 'Total Damage Dealt'
			sheet.cell(row = 5, column = 1).value = 'Total Damage Taken'

			for i in range(3,6,2):
				for j in range(1,14):
					sheet.cell(row = i, column = j).fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
			
			for i in range(6,21):
				for j in range(1,14):
					sheet.cell(row = i, column = j).fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

			for i in range(22,323,3):
				for j in range(1,14):
					sheet.cell(row = i, column = j).fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
			
			for i in range(23,324,3):
				for j in range(1,14):
					sheet.cell(row = i, column = j).fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
		
			for i in range(24,325,3):
				for j in range(1,14):
					sheet.cell(row = i, column = j).fill = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
			
			for letter in ['A','B','C','D','E','F','G','H','I','J','K','L','M']:
				sheet.column_dimensions[letter].width = 24
		#If member is already in sheet	
		else:
			sheet = book[member_dict.get(str(member), str(member))]

		start_row = 22
		for row in range(22,323,3):
			if sheet.cell(row = row, column = 1).value == None:
				start_row = row
				break

		sheet.cell(row = start_row, column = 1).value = f'{datetime.date(datetime.now())} Net Damage:'
		sheet.cell(row = start_row+1, column = 1).value = f'{datetime.date(datetime.now())} Damage Dealt:'
		sheet.cell(row = start_row+2, column = 1).value = f'{datetime.date(datetime.now())} Damage Taken:'

		#Calculates the proper damage deal in money value
		dealt = calc_stats(member_records[member][0], rss_prices)
		taken = calc_stats(member_records[member][1], rss_prices)

		#Updates the values below in the date section
		for col, key in enumerate(['Infrastructure','Soldiers','Tanks','Planes','Ships','Missiles','Nukes','All Units', 'Gas and Munition', 'Money','Looted','Total']):
			sheet.cell(row = start_row+1, column = col+2).value = dealt[key]
			sheet.cell(row = start_row+2, column = col+2).value = taken[key]

		#Updates the total damage dealt and taken
		for index in range(2,14):
			dealt = sheet.cell(row = start_row+1, column = index).value
			taken = sheet.cell(row = start_row+2, column = index).value
			if '(' in dealt:
				d_units, d_cost = dealt.split(' ($')
				d_cost = d_cost.replace(')', '')
				t_units, t_cost = taken.split(' ($')
				t_cost = t_cost.replace(')', '')
				sheet.cell(row = start_row, column = index).value = f'{round(float(d_units) - float(t_units),2)} (${round(float(d_cost) - float(t_cost),2)})'
			else:
				taken = float(taken.replace('$', ''))
				dealt = float(dealt.replace('$', ''))
				sheet.cell(row = start_row, column = index).value = f'${dealt - taken}'

		#dataframe time!
		df_net = create_df(sheet, 0, start_row+3)
		df_dealt = create_df(sheet, 1, start_row+3)
		df_taken = create_df(sheet, 2, start_row+3)
		
		#Deletes old graphs
		for index in range(len(sheet._images)-1,-1, -1):
			del sheet._images[index]

		#Creates new graphs
		total_dealt = df_dealt[['cinfra', 'csold','ctanks', 'cair', 'cships', 'cmiss', 'cnuke','gm', 'mon_des', 'loot']].sum()
		pie_graph = plt.figure(figsize = (5,2.9))
		print(len(total_dealt))
		plt.pie(total_dealt)
		pie_graph.legend(['Infrastructure','Soldiers','Tanks','Planes','Ships','Missiles','Nukes','Gas and Munition','Money','Looted'], title = 'Damage Dealt')
		pie_graph.savefig(f"Graphs/Pie Dealt {member}.png")
		img = openpyxl.drawing.image.Image(f'Graphs/Pie Dealt {member}.png')
		img.anchor = 'E6'
		sheet.add_image(img)

		total_taken = df_taken[['cinfra', 'csold','ctanks', 'cair', 'cships', 'cmiss', 'cnuke','gm', 'mon_des', 'loot']].sum()
		pie_graph = plt.figure(figsize = (5,2.9))
		print(len(total_taken))
		plt.pie(total_taken)
		pie_graph.legend(['Infrastructure','Soldiers','Tanks','Planes','Ships','Missiles','Nukes','Gas and Munition','Money','Looted'], title = 'Damage Taken')
		pie_graph.savefig(f"Graphs/Pie Taken {member}.png")
		img = openpyxl.drawing.image.Image(f'Graphs/Pie Taken {member}.png')
		img.anchor = 'H6'
		sheet.add_image(img)
		print(sheet)
		line = plt.figure()
		sns.lineplot(data=df_net, x='date', y='total')
		sns.lineplot(data=df_net, x='date', y='csold')
		sns.lineplot(data=df_net, x='date', y='ctanks')
		sns.lineplot(data=df_net, x='date', y='cair')
		sns.lineplot(data=df_net, x='date', y='cships')
		sns.lineplot(data=df_net, x='date', y='cmiss')
		sns.lineplot(data=df_net, x='date', y='cnuke')
		sns.lineplot(data=df_net, x='date', y='gm')
		sns.lineplot(data=df_net, x='date', y='mon_des')
		sns.lineplot(data=df_net, x='date', y='loot')
		line.set_size_inches(5, 2.9)
		line.savefig(f"Graphs/NetLine {member}.png")
		img = openpyxl.drawing.image.Image(f'Graphs/NetLine {member}.png')
		img.anchor = 'B6'
		sheet.add_image(img)
		for i, df in enumerate([df_net, df_dealt, df_taken]):
			for j, val in enumerate(['infra', 'sold', 'tanks', 'air', 'ships','miss', 'nuke']):
				sheet.cell(row = 3+i, column = j+2).value = f'{df[val].sum()} (${df[f"c{val}"].sum()})'
			for j, val in enumerate(['all', 'gm', 'mon_des', 'loot', 'total']):
				sheet.cell(row = 3+i, column = j+9).value = f'${df[val].sum()}'
		plt.close('all')
	book.save('Stat Tracker.xlsx')


def calc_stats(damage, rss_prices):
	'''
	Calculates the damage in $ values
	:param damage: the pure damage in units
	:param rss_prices: price of resources
	:return: the format to fill in spreadsheet 'units ($unit value)'
	'''
	stats ={'Infrastructure': '','Soldiers': '','Tanks': '','Planes': '','Ships': '','Missiles': '',
	'Nukes': '','All Units': '', 'Gas and Munition': '', 'Money': '','Looted': '','Total': ''}

	stats['Infrastructure'] = f'{round(damage["infra_destroyed"], 2)} (${round(damage["infra_destroyed_value"], 2)})'

	soldier_cost = damage['sold'] * 5
	stats['Soldiers'] = f'{damage["sold"]} (${soldier_cost})'

	tank_cost = damage['tanks'] * (60+0.5*rss_prices['Steel'])
	stats['Tanks'] = f'{damage["tanks"]} (${tank_cost})'

	plane_cost = damage['air'] * (4000+5*rss_prices['Aluminum'])
	stats['Planes'] = f'{damage["air"]} (${plane_cost})'

	ship_cost = damage['ships'] * (50000+25*rss_prices['Steel'])
	stats['Ships'] = f'{damage["ships"]} (${ship_cost})'

	miss_cost = damage['miss'] * (150000+500*rss_prices['Gasoline']+250*rss_prices['Uranium']+750*rss_prices['Aluminum'])
	stats['Missiles'] = f'{damage["miss"]} (${miss_cost})'

	nuke_cost = damage['nuke'] * (150000+75*rss_prices['Gasoline']+75*rss_prices['Munitions']+100*rss_prices['Aluminum'])
	stats['Nukes'] = f'{damage["nuke"]} (${nuke_cost})'

	all_cost = soldier_cost + tank_cost + plane_cost + ship_cost + miss_cost + nuke_cost
	stats['All Units'] = f'${round(all_cost, 2)}'

	stats['Money'] = f'${round(damage["mon_destroyed"], 2)}'

	gas_mun_cost = round((damage['gas'] * rss_prices['Gasoline'] + damage['mun'] * rss_prices['Munitions']), 2)
	stats['Gas and Munition'] = f'${gas_mun_cost}'
	
	looted = damage['beige_loot'][0] * rss_prices['Coal'] + damage['beige_loot'][1] * rss_prices['Oil'] +\
	damage['beige_loot'][2] * rss_prices['Uranium'] + damage['beige_loot'][3] * rss_prices['Iron'] +\
	damage['beige_loot'][4] * rss_prices['Bauxite'] + damage['beige_loot'][5] * rss_prices['Lead'] +\
	damage['beige_loot'][6] * rss_prices['Gasoline'] + damage['beige_loot'][7] * rss_prices['Munitions'] +\
	damage['beige_loot'][8] * rss_prices['Steel'] + damage['beige_loot'][9] * rss_prices['Aluminum'] +\
	damage['beige_loot'][10] * rss_prices['Food'] + damage['money_looted']

	stats['Looted'] = f'${looted}'
	stats['Total'] = f'${all_cost + gas_mun_cost + round(damage["mon_destroyed"],2) + looted + round(damage["infra_destroyed_value"], 2)}'

	return stats



def create_df(sheet, x, start_row):
	'''
	Turns spreadsheet into a dataframe
	:return: dataframe of spreadsheet values with all the dates
	'''
	df = pd.DataFrame(columns = ['date','infra', 'sold', 'tanks', 'air', 'ships',
		'miss', 'nuke', 'all', 'gm', 'mon_des', 'loot', 'total', 'cinfra', 'csold',
		'ctanks', 'cair', 'cships', 'cmiss', 'cnuke'], dtype = object)
	for i in range(22+x, start_row+x, 3):
		row = [sheet.cell(row = i, column = j).value for j in range(1,14)]
		row[0] = datetime.strptime(row[0].split(' ')[0], '%Y-%m-%d').date()
		for index, value in enumerate(row[1:]):
			if '(' in value:
				units, cost = row[index+1].split(' ($')
				cost = float(cost.replace(')', ''))
				units = float(units)
				row[index+1] = units
				row.append(cost)
			else:
				row[index+1] = float(row[index+1].replace('$', ''))
		df.loc[len(df.index)] = row
	return df


if __name__ == '__main__':
	#last_war_save = pd.read_excel('Last Active Wars.xlsx')
	#load_wars()
	wars = [736361,736346,736342,736272]
	member_records = {}
	records = calc_war(48730, 736367)
	member_records['EvilPiggyFooFoo'] = records[0:2]
	for war in wars:
		records = calc_war(48730, war)
		for key in member_records['EvilPiggyFooFoo'][0]:
			member_records['EvilPiggyFooFoo'][0][key] += records[0][key]
		for key in member_records['EvilPiggyFooFoo'][1]:
			member_records['EvilPiggyFooFoo'][1][key] += records[1][key]

	print(member_records)

	rss_prices = {}
	#Gets current rss prices
	for resource in ['Coal', 'Oil', 'Uranium', 'Iron', 'Bauxite', 'Lead', 'Gasoline', 'Munitions', 'Steel', 'Aluminum', 'Food']:
		rss_json = war_json = requests.get(f'https://politicsandwar.com/api/tradeprice/resource={resource.lower()}&key={API_KEY}').json()
		rss_prices[resource] = int(rss_json['avgprice'])
	print(calc_stats(member_records['EvilPiggyFooFoo'][1], rss_prices))